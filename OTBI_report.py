# from fastapi import FastAPI, Query, HTTPException
# from fastapi.responses import Response
# import base64
# import requests
# import xml.etree.ElementTree as ET

# # Config
# from config import FUSION_BASE_URL, FUSION_USERNAME, FUSION_PASSWORD

# app = FastAPI(title="Fusion BI Report API")


# # -----------------------------
# # Helpers
# # -----------------------------
# def extract_report_bytes(response_text: str):
#     try:
#         root = ET.fromstring(response_text)
#         ns = {"ns2": "http://xmlns.oracle.com/oxp/service/PublicReportService"}
#         return root.find(".//ns2:reportBytes", ns).text
#     except Exception as e:
#         print("Extraction error:", e)
#         return None


# def extract_fault(response_text: str):
#     """Extract SOAP fault message"""
#     try:
#         root = ET.fromstring(response_text)
#         return root.find(".//{http://www.w3.org/2003/05/soap-envelope}Text").text
#     except:
#         return "Unknown Fusion error"


# def call_fusion(reportXDOpath: str, format: str):
#     """Call Fusion BI Publisher SOAP API"""

#     url = f"{FUSION_BASE_URL}/xmlpserver/services/ExternalReportWSSService"

#     soap_body = f"""
# <soapenv:Envelope xmlns:soapenv="http://www.w3.org/2003/05/soap-envelope"
#                   xmlns:pub="http://xmlns.oracle.com/oxp/service/PublicReportService">
#    <soapenv:Header/>
#    <soapenv:Body>
#       <pub:runReport>
#          <pub:reportRequest>
#             <pub:attributeFormat>{format}</pub:attributeFormat>
#             <pub:attributeLocale>en-US</pub:attributeLocale>
#             <pub:reportAbsolutePath>{reportXDOpath}</pub:reportAbsolutePath>
#             <pub:sizeOfDataChunkDownload>-1</pub:sizeOfDataChunkDownload>
#             <pub:flattenXML>true</pub:flattenXML>
#          </pub:reportRequest>
#       </pub:runReport>
#    </soapenv:Body>
# </soapenv:Envelope>
# """

#     headers = {
#         "Content-Type": "application/soap+xml; charset=utf-8"
#     }

#     response = requests.post(
#         url,
#         data=soap_body,
#         headers=headers,
#         auth=(FUSION_USERNAME, FUSION_PASSWORD)
#     )

#     return response


# # -----------------------------
# # API
# # -----------------------------
# @app.get("/getReport")
# def get_report(
#     reportXDOpath: str = Query(..., description="Report Path (e.g. /JOHN/John_report.xdo)"),
#     format: str = Query("xlsx", description="Format: xlsx, pdf, html, etc.")
# ):
#     try:
#         # 🔥 Call Fusion
#         response = call_fusion(reportXDOpath, format.lower())

#         print("STATUS:", response.status_code)
#         print("RESPONSE:", response.text[:1000])

#         # ❌ Handle HTTP error
#         if response.status_code != 200:
#             fault = extract_fault(response.text)
#             raise HTTPException(status_code=500, detail=fault)

#         # ❌ Extract report data
#         report_b64 = extract_report_bytes(response.text)

#         if not report_b64:
#             fault = extract_fault(response.text)
#             raise HTTPException(status_code=500, detail=fault)

#         import base64

#         report_bytes = base64.b64decode(report_b64.strip())
#         print("Decoded bytes size:", len(report_bytes))

#         # =========================
#         # ✅ FORMAT HANDLING
#         # =========================

#         # Excel
#         if format.lower() == "xlsx":
#             print("Returning Excel file...")
#             return Response(
#                 content=report_bytes,
#                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                 headers={
#                     "Content-Disposition": "attachment; filename=report.xlsx"
#                 }
#             )

#         # PDF
#         elif format.lower() == "pdf":
#             return Response(
#                 content=report_bytes,
#                 media_type="application/pdf",
#                 headers={
#                     "Content-Disposition": "attachment; filename=report.pdf"
#                 }
#             )

#         # HTML
#         elif format.lower() == "html":
#             return Response(
#                 content=report_bytes,
#                 media_type="text/html"
#             )

#         # PowerPoint
#         elif format.lower() == "pptx":
#             return Response(
#                 content=report_bytes,
#                 media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
#                 headers={
#                     "Content-Disposition": "attachment; filename=report.pptx"
#                 }
#             )

#         # RTF
#         elif format.lower() == "rtf":
#             return Response(
#                 content=report_bytes,
#                 media_type="application/rtf",
#                 headers={
#                     "Content-Disposition": "attachment; filename=report.rtf"
#                 }
#             )

#         # ❗ Unsupported format fallback
#         else:
#             return {
#                 "format": format,
#                 "message": "Format not explicitly handled, returning raw size",
#                 "file_size_bytes": len(report_bytes)
#             }

#     except Exception as e:
#         raise HTTPException(status_code=500, detail=str(e))


# # -----------------------------
# # Entry Point
# # -----------------------------
# if __name__ == "__main__":
#     import uvicorn
#     uvicorn.run("OTBI_report:app", host="0.0.0.0", port=8000, reload=True)



from fastapi import FastAPI, Query, HTTPException
import base64
import requests
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from io import BytesIO

# Config
from config import FUSION_BASE_URL, FUSION_USERNAME, FUSION_PASSWORD

app = FastAPI(title="Fusion BI Report API (Excel → JSON)")


# -----------------------------
# Helpers
# -----------------------------
def extract_report_bytes(response_text: str):
    try:
        root = ET.fromstring(response_text)
        ns = {"ns2": "http://xmlns.oracle.com/oxp/service/PublicReportService"}
        return root.find(".//ns2:reportBytes", ns).text
    except:
        return None


def extract_fault(response_text: str):
    try:
        root = ET.fromstring(response_text)
        return root.find(".//{http://www.w3.org/2003/05/soap-envelope}Text").text
    except:
        return "Unknown Fusion error"


def call_fusion(reportXDOpath: str):
    url = f"{FUSION_BASE_URL}/xmlpserver/services/ExternalReportWSSService"

    soap_body = f"""
<soapenv:Envelope xmlns:soapenv="http://www.w3.org/2003/05/soap-envelope"
                  xmlns:pub="http://xmlns.oracle.com/oxp/service/PublicReportService">
   <soapenv:Header/>
   <soapenv:Body>
      <pub:runReport>
         <pub:reportRequest>
            <pub:attributeFormat>xlsx</pub:attributeFormat>
            <pub:attributeLocale>en-US</pub:attributeLocale>
            <pub:reportAbsolutePath>{reportXDOpath}</pub:reportAbsolutePath>
            <pub:sizeOfDataChunkDownload>-1</pub:sizeOfDataChunkDownload>
            <pub:flattenXML>true</pub:flattenXML>
            <pub:byPassCache>true</pub:byPassCache>
         </pub:reportRequest>
      </pub:runReport>
   </soapenv:Body>
</soapenv:Envelope>
"""

    headers = {
        "Content-Type": "application/soap+xml; charset=utf-8"
    }

    response = requests.post(
        url,
        data=soap_body,
        headers=headers,
        auth=(FUSION_USERNAME, FUSION_PASSWORD)
    )

    return response


# -----------------------------
# Excel → JSON Converter
# -----------------------------
def excel_to_json(excel_bytes):
    wb = load_workbook(BytesIO(excel_bytes), data_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return []

    # 🔥 Step 1: Detect header row
    header_row_index = None
    for i, row in enumerate(rows):
        non_empty_cells = [cell for cell in row if cell is not None]

        if len(non_empty_cells) >= 3 and all(isinstance(cell, str) for cell in non_empty_cells):
            header_row_index = i
            break

    if header_row_index is None:
        raise Exception("Header row not found")

    raw_headers = rows[header_row_index]

    # 🔥 Step 2: Clean headers
    headers = []
    valid_indexes = []

    for i, h in enumerate(raw_headers):
        if h and str(h).strip():
            clean = str(h).replace('"', '').split('.')[-1].strip()

            # Skip completely useless headers
            if clean.lower() not in ["", "none"]:
                headers.append(clean)
                valid_indexes.append(i)

    # 🔥 Step 3: Extract only valid columns
    data = []
    for row in rows[header_row_index + 1:]:
        if not any(row):
            continue

        record = {}
        for idx, col_index in enumerate(valid_indexes):
            record[headers[idx]] = row[col_index] if col_index < len(row) else None

        data.append(record)

    return data


# -----------------------------
# API
# -----------------------------
@app.get("/getReport")
def get_report(
    reportXDOpath: str = Query(..., description="Report Path (e.g. /JOHN/John_report.xdo)")
):
    try:
        # 🔥 Call Fusion
        response = call_fusion(reportXDOpath)

        print("STATUS:", response.status_code)

        if response.status_code != 200:
            fault = extract_fault(response.text)
            raise HTTPException(status_code=500, detail=fault)

        report_b64 = extract_report_bytes(response.text)

        if not report_b64:
            fault = extract_fault(response.text)
            raise HTTPException(status_code=500, detail=fault)

        report_bytes = base64.b64decode(report_b64.strip())

        print("Excel size:", len(report_bytes))

        # 🔥 Convert Excel → JSON
        json_data = excel_to_json(report_bytes)

        return {
            "data": json_data,
            "count": len(json_data)
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# -----------------------------
# Entry Point
# -----------------------------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("OTBI_report:app", host="0.0.0.0", port=8000, reload=True)